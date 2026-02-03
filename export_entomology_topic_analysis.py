# -*- coding: utf-8 -*-
"""
수목해충학 기출문제 주제별 분석 엑셀 내보내기
"""
import os
import sys
import django
import re

sys.path.append('.')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from exam.models import Question, Subject
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

subject = Subject.objects.get(name__contains='수목해충학')

# 주제 분류 함수
topic_priority = [
    ('살충제/약제', [r'살충제', r'농약', r'약제', r'네오니코티노이드', r'유기인계', r'카바메이트', r'피레스로이드', r'BT제', r'훈증', r'살비제', r'약해', r'잔류']),
    ('방제법', [r'방제', r'생물적', r'화학적', r'물리적', r'천적', r'IPM', r'종합.*관리', r'페로몬']),
    ('예찰/트랩', [r'예찰', r'트랩', r'페로몬트랩', r'유아등', r'피해율', r'밀도조사']),
    ('곤충 형태', [r'머리', r'가슴', r'복부', r'배$', r'다리', r'날개', r'더듬이', r'겹눈', r'홑눈', r'입틀', r'큰턱', r'표피', r'외골격', r'감각기', r'기문', r'순환', r'호흡', r'신경']),
    ('곤충 생리', [r'소화', r'기관계', r'말피기관', r'혈림프', r'호르몬', r'내분비', r'탈피호르몬', r'유약호르몬', r'배설']),
    ('생활사/변태', [r'변태', r'완전변태', r'불완전변태', r'탈피', r'우화', r'산란', r'생활사', r'발육', r'세대', r'알$', r'유충', r'번데기', r'성충', r'령', r'월동', r'휴면', r'적산온도']),
    ('나방류', [r'나방', r'명나방', r'솔나방', r'미국흰불나방', r'매미나방', r'독나방', r'재주나방', r'밤나방', r'잎말이']),
    ('딱정벌레류', [r'하늘소', r'바구미', r'나무좀', r'풍뎅이', r'딱정벌레', r'잎벌레']),
    ('노린재/진딧물', [r'노린재', r'진딧물', r'깍지벌레', r'매미충', r'선녀벌레', r'꽃매미', r'매미$', r'감로']),
    ('파리/벌류', [r'파리', r'혹파리', r'솔잎혹파리', r'잎벌']),
    ('응애류', [r'응애', r'진드기']),
    ('해충 피해', [r'식엽성', r'천공성', r'흡즙성', r'충영', r'종실', r'구과', r'가해', r'피해']),
    ('분류/동정', [r'분류', r'동정', r'학명', r'곤충강', r'절지동물', r'목\)']),
]

def classify_single(q):
    full_text = f"{q.content} {q.choice1} {q.choice2} {q.choice3} {q.choice4} {q.choice5}"
    for topic, patterns in topic_priority:
        for pattern in patterns:
            if re.search(pattern, full_text, re.IGNORECASE):
                return topic
    return '기타'

# 워크북 생성
wb = openpyxl.Workbook()

# 스타일 정의
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 주제별 색상
topic_colors = {
    '곤충 형태': 'C6EFCE',
    '곤충 생리': 'C6EFCE',
    '생활사/변태': 'FFEB9C',
    '방제법': 'BDD7EE',
    '살충제/약제': 'BDD7EE',
    '예찰/트랩': 'BDD7EE',
    '나방류': 'F8CBAD',
    '딱정벌레류': 'F8CBAD',
    '노린재/진딧물': 'F8CBAD',
    '파리/벌류': 'F8CBAD',
    '응애류': 'F8CBAD',
    '해충 피해': 'D9D9D9',
    '분류/동정': 'D9D9D9',
    '기타': 'D9D9D9',
}

# 각 회차별 시트 생성
for round_num in range(5, 12):
    questions = Question.objects.filter(subject=subject, exam__round_number=round_num).order_by('number')

    if not questions.exists():
        continue

    # 시트 생성
    if round_num == 5:
        ws = wb.active
        ws.title = f'{round_num}회차'
    else:
        ws = wb.create_sheet(title=f'{round_num}회차')

    # 헤더
    headers = ['번호', '주제', '문제 내용']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # 데이터
    for row, q in enumerate(questions, 2):
        topic = classify_single(q)
        content_short = q.content[:80].replace('\n', ' ').strip()
        if len(q.content) > 80:
            content_short += '...'

        # 번호
        cell = ws.cell(row=row, column=1, value=q.number)
        cell.border = thin_border
        cell.alignment = center_align

        # 주제
        cell = ws.cell(row=row, column=2, value=topic)
        cell.border = thin_border
        cell.alignment = center_align
        if topic in topic_colors:
            cell.fill = PatternFill(start_color=topic_colors[topic], end_color=topic_colors[topic], fill_type='solid')

        # 문제 내용
        cell = ws.cell(row=row, column=3, value=content_short)
        cell.border = thin_border
        cell.alignment = left_align

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 80

# 요약 시트 생성
ws_summary = wb.create_sheet(title='회차별 요약')

# 요약 헤더
summary_headers = ['회차', '곤충형태', '곤충생리', '생활사/변태', '방제법', '살충제/약제', '예찰/트랩', '나방류', '딱정벌레류', '노린재/진딧물', '파리/벌류', '기타', '합계']
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

# 요약 데이터
topic_list = ['곤충 형태', '곤충 생리', '생활사/변태', '방제법', '살충제/약제', '예찰/트랩', '나방류', '딱정벌레류', '노린재/진딧물', '파리/벌류', '기타']

for row, round_num in enumerate(range(5, 12), 2):
    questions = Question.objects.filter(subject=subject, exam__round_number=round_num)

    # 회차
    cell = ws_summary.cell(row=row, column=1, value=f'{round_num}회')
    cell.border = thin_border
    cell.alignment = center_align

    # 주제별 카운트
    topic_counts = {}
    for q in questions:
        topic = classify_single(q)
        topic_counts[topic] = topic_counts.get(topic, 0) + 1

    total = 0
    for col, topic in enumerate(topic_list, 2):
        count = topic_counts.get(topic, 0)
        total += count
        cell = ws_summary.cell(row=row, column=col, value=count if count > 0 else '')
        cell.border = thin_border
        cell.alignment = center_align

    # 합계
    cell = ws_summary.cell(row=row, column=len(summary_headers), value=total)
    cell.border = thin_border
    cell.alignment = center_align
    cell.font = Font(bold=True)

# 합계 행
row = 9
cell = ws_summary.cell(row=row, column=1, value='합계')
cell.font = Font(bold=True)
cell.border = thin_border
cell.alignment = center_align

for col in range(2, len(summary_headers) + 1):
    total = sum(ws_summary.cell(row=r, column=col).value or 0 for r in range(2, 9))
    cell = ws_summary.cell(row=row, column=col, value=total)
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.alignment = center_align

# 열 너비 조정
for col in range(1, len(summary_headers) + 1):
    ws_summary.column_dimensions[get_column_letter(col)].width = 12

# 저장
output_path = '수목해충학_기출문제_주제분석.xlsx'
wb.save(output_path)
print(f'엑셀 파일 생성 완료: {output_path}')
