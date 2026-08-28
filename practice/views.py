from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db.models import Q
import openpyxl
from .models import Book, Chapter, PracticeQuestion, PracticeAttempt, ChapterContent, ChapterPost


@login_required
def book_list(request):
    """교재 목록"""
    books = Book.objects.all()
    return render(request, 'practice/book_list.html', {'books': books})


def _exam_star_map(book):
    """목차별 기출 빈도를 용어집 화면과 같은 모양으로 만든다.

    막대 5칸 + 문항 수 + 회차 배지. 목차에 이어 붙인 주제별 문제집을 센다.
    목차마다 따로 읽으면 느리므로 한 번에 가져와 사전으로 만든다.
    """
    from exam.models import TopicQuestionSetItem

    rows = (TopicQuestionSetItem.objects
            .filter(question_set__chapter__book=book,
                    question_set__chapter__isnull=False)
            .values_list('question_set__chapter_id',
                         'question__exam__round_number',
                         'question__number'))

    by_chapter = {}
    for cid, rnd, num in rows:
        if rnd and num:
            by_chapter.setdefault(cid, set()).add((rnd, num))

    out = {}
    for cid, pairs in by_chapter.items():
        items = sorted(pairs)
        out[cid] = {
            'n': len(items),
            # 회차-문제번호 (5-25 꼴)
            'refs': [{'round': r, 'label': '%d-%d' % (r, q)} for r, q in items],
        }
    return out


@login_required
def chapter_list(request, book_id):
    """교재의 목차 트리"""
    book = get_object_or_404(Book, id=book_id)

    # 최상위 목차만 가져오기 (하위는 템플릿에서 재귀적으로)
    root_chapters = Chapter.objects.filter(book=book, parent=None).order_by('order', 'code')

    return render(request, 'practice/chapter_list.html', {
        'book': book,
        'root_chapters': root_chapters,
        'exam_stars': _exam_star_map(book),
    })


@login_required  
def practice_questions(request, chapter_id):
    """목차별 문제 풀이"""
    chapter = get_object_or_404(Chapter, id=chapter_id)
    questions = PracticeQuestion.objects.filter(chapter=chapter).order_by('number')
    
    # 사용자의 이전 풀이 기록 가져오기
    user_attempts = {}
    if request.user.is_authenticated:
        attempts = PracticeAttempt.objects.filter(
            user=request.user,
            question__in=questions
        ).order_by('-attempted_at')
        
        # 각 문제별 최신 시도만 저장
        for attempt in attempts:
            if attempt.question_id not in user_attempts:
                user_attempts[attempt.question_id] = attempt
    
    return render(request, 'practice/practice_questions.html', {
        'chapter': chapter,
        'questions': questions,
        'user_attempts': user_attempts,
    })


@login_required
def submit_answer(request, question_id):
    """정답 제출 (AJAX)"""
    if request.method == 'POST':
        question = get_object_or_404(PracticeQuestion, id=question_id)
        
        try:
            selected = int(request.POST.get('answer', 0))
        except ValueError:
            return JsonResponse({'error': 'Invalid answer'}, status=400)
        
        is_correct = (selected == question.answer)
        
        # 풀이 기록 저장
        PracticeAttempt.objects.create(
            user=request.user,
            question=question,
            selected_answer=selected,
            is_correct=is_correct
        )
        
        return JsonResponse({
            'is_correct': is_correct,
            'correct_answer': question.answer,
            'explanation': question.explanation,
        })
    
    return JsonResponse({'error': 'Invalid method'}, status=405)


@login_required
def upload_questions(request):
    """엑셀 파일로 문제 일괄 업로드"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 접근 가능합니다.')
        return redirect('practice:book_list')
    
    books = Book.objects.all()
    
    if request.method == 'POST':
        book_id = request.POST.get('book_id')
        file = request.FILES.get('excel_file')
        
        if not book_id or not file:
            messages.error(request, '교재와 파일을 모두 선택해주세요.')
            return redirect('practice:upload_questions')
        
        book = get_object_or_404(Book, id=book_id)
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            created_count = 0
            error_rows = []
            
            # 첫 행은 헤더로 가정
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # 빈 행 스킵
                    continue
                
                chapter_code = str(row[0]).strip()
                content = str(row[1]).strip() if row[1] else ''
                choice1 = str(row[2]).strip() if row[2] else ''
                choice2 = str(row[3]).strip() if row[3] else ''
                choice3 = str(row[4]).strip() if row[4] else ''
                choice4 = str(row[5]).strip() if row[5] else ''
                choice5 = str(row[6]).strip() if row[6] else ''
                
                try:
                    answer = int(row[7]) if row[7] else 0
                except:
                    answer = 0
                    
                explanation = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                
                # 목차 찾기
                try:
                    chapter = Chapter.objects.get(book=book, code=chapter_code)
                except Chapter.DoesNotExist:
                    error_rows.append(f"행 {row_idx}: 목차 '{chapter_code}'를 찾을 수 없음")
                    continue
                
                # 문제 번호 자동 할당
                last_q = PracticeQuestion.objects.filter(chapter=chapter).order_by('-number').first()
                next_number = (last_q.number + 1) if last_q else 1
                
                PracticeQuestion.objects.create(
                    chapter=chapter,
                    number=next_number,
                    content=content,
                    choice1=choice1,
                    choice2=choice2,
                    choice3=choice3,
                    choice4=choice4,
                    choice5=choice5,
                    answer=answer,
                    explanation=explanation,
                )
                created_count += 1
            
            if error_rows:
                messages.warning(request, f'{created_count}개 문제 추가됨. 오류: {"; ".join(error_rows[:5])}')
            else:
                messages.success(request, f'{created_count}개 문제가 추가되었습니다.')
                
        except Exception as e:
            messages.error(request, f'파일 처리 중 오류: {str(e)}')
        
        return redirect('practice:upload_questions')
    
    return render(request, 'practice/upload.html', {'books': books})


@login_required
def upload_chapters(request):
    """엑셀 파일로 목차 일괄 업로드"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 접근 가능합니다.')
        return redirect('practice:book_list')
    
    books = Book.objects.all()
    
    if request.method == 'POST':
        book_id = request.POST.get('book_id')
        file = request.FILES.get('excel_file')
        
        if not book_id or not file:
            messages.error(request, '교재와 파일을 모두 선택해주세요.')
            return redirect('practice:upload_chapters')
        
        book = get_object_or_404(Book, id=book_id)
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            
            # 첫 행은 헤더로 가정
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # 빈 행 스킵
                    continue
                
                code = str(row[0]).strip()
                title = str(row[1]).strip() if row[1] else ''
                
                # level 자동 계산 (점 개수 + 1)
                level = code.count('.') + 1
                
                # parent 찾기 (상위 목차)
                parent = None
                if '.' in code:
                    parent_code = '.'.join(code.split('.')[:-1])
                    try:
                        parent = Chapter.objects.get(book=book, code=parent_code)
                    except Chapter.DoesNotExist:
                        pass  # 상위가 없으면 루트로
                
                # order 자동 할당
                order = row_idx
                
                # 기존 목차가 있으면 업데이트, 없으면 생성
                chapter, created = Chapter.objects.update_or_create(
                    book=book,
                    code=code,
                    defaults={
                        'title': title,
                        'level': level,
                        'parent': parent,
                        'order': order,
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
            
            messages.success(request, f'목차 {created_count}개 생성, {updated_count}개 업데이트되었습니다.')
                
        except Exception as e:
            messages.error(request, f'파일 처리 중 오류: {str(e)}')
        
        return redirect('practice:upload_chapters')
    
    return render(request, 'practice/upload_chapters.html', {'books': books})


def _qa_body(html):
    """관련 Q&A 본문을 화면에 맞게 다듬는다.

    게시글은 HTML 로 저장되는데 그 안에 마크다운 글머리표(*)와
    LaTeX 수식($...$)이 섞여 들어온 것이 있다. 그대로 두면
    '* 보통 1년에' '$15^\\circ\\text{C}$' 처럼 원문이 노출된다.
    """
    import re

    if not html:
        return ''

    text = str(html)

    # 수식일 이유가 없는 것은 글자로 되돌린다 (용어 화면과 같은 처리)
    try:
        from glossary.views import _unwrap_plain_math
        text = _unwrap_plain_math(text)
    except Exception:
        pass

    # $15^\circ\text{C}$ -> 15°C
    def _tex(m):
        inner = m.group(1)
        # ^\circ 는 '도' 표기다. ^ 와 중괄호까지 함께 없애야
        # 아래 '진짜 수식' 검사(_ ^ 백슬래시)에 걸리지 않는다
        inner = re.sub(r'\^\s*\{?\s*\\+circ\s*\}?', '°', inner)
        inner = re.sub(r'\\+circ', '°', inner)
        inner = re.sub(r'\\+sim|\\+thicksim', '~', inner)
        inner = re.sub(r'\\+times', '×', inner)
        inner = re.sub(r'\\+text\{([^{}]*)\}', lambda x: x.group(1), inner)
        inner = re.sub(r'\\+%', '%', inner)
        inner = inner.replace('\\ ', ' ').replace('{', '').replace('}', '')
        if re.search(r'[\\_^]', inner):
            return m.group(0)          # 진짜 수식은 그대로 둔다
        return inner

    text = re.sub(r'\$([^$\n]{1,40})\$', _tex, text)

    # 줄 앞의 '* ' 를 글머리표로 (HTML 안이라 마크다운이 돌지 않았다)
    text = re.sub(r'(?m)(^|<br\s*/?>\s*\n?)\s*\*\s+', r'\1• ', text)

    return text


def _qa_preview(html, limit=180):
    """관련 Q&A 미리보기. 다듬은 뒤 앞부분만 남긴다."""
    from mypage.views import _plain_preview

    return _plain_preview(_qa_body(html), limit)


@login_required
@require_POST
def api_post_hit(request, post_id):
    """관련 Q&A 를 펼쳤을 때 조회수를 올린다.

    본래 제목을 눌러 게시글 화면으로 가면 그곳에서 올랐는데,
    그 자리에서 펼치도록 바꾸면서 조회가 잡히지 않았다.
    게시판(bbs)과 마찬가지로 펼칠 때마다 센다.
    """
    from bbs.models import Post
    from django.db.models import F

    updated = Post.objects.filter(id=post_id).update(hits=F('hits') + 1)
    if not updated:
        return JsonResponse({'ok': False}, status=404)

    hits = Post.objects.filter(id=post_id).values_list('hits', flat=True).first()
    return JsonResponse({'ok': True, 'counted': True, 'hits': hits})


def _chapter_exam_sets(chapter):
    """목차에 이어 붙인 주제별 문제집을 회차별 요약과 함께 돌려준다.

    화면에는 '5회 3문제' 처럼 배지로 보여 주고,
    '문제 풀기'를 누르면 그 문제집 전체를 푼다.
    """
    from exam.models import TopicQuestionSet

    sets = (TopicQuestionSet.objects
            .filter(chapter=chapter, is_public=True)
            .prefetch_related('items__question__exam')
            .order_by('order', '-created_at'))

    rows = []
    for ts in sets:
        questions = []
        for item in ts.items.all().order_by('order'):
            q = item.question
            exam = getattr(q, 'exam', None)
            if exam is None:
                continue
            questions.append({
                # 배지는 '5-1' 처럼 회차-문제번호
                'badge': '%s-%s' % (exam.round_number, q.number),
                'round': exam.round_number,   # 배지 색을 회차별로 주려고
                'content': q.content,
            })
        if not questions:
            continue
        rows.append({
            'set': ts,
            'total': len(questions),
            'questions': questions,
        })
    return rows


@login_required
def chapter_detail(request, chapter_id):
    """목차 컨텐츠 상세 보기"""
    chapter = get_object_or_404(Chapter, id=chapter_id)
    
    # 컨텐츠가 있는지 확인
    try:
        content = chapter.content
    except ChapterContent.DoesNotExist:
        content = None
    
    # 관련 문제 가져오기
    questions = PracticeQuestion.objects.filter(chapter=chapter).order_by('number')
    
    # Natural sort function for code like "1.2.3"
    def natural_sort_key(ch):
        if not ch.code:
            return (999,)
        parts = ch.code.split('.')
        result = []
        for p in parts:
            try:
                result.append(int(p))
            except ValueError:
                result.append(999)
        return tuple(result)
    
    # 현재 목차 기준으로 앞뒤 5개씩 가져오기 (레벨 무관, code 기준 정렬)
    all_chapters = list(Chapter.objects.filter(book=chapter.book))
    all_chapters.sort(key=natural_sort_key)
    
    # 현재 목차의 인덱스 찾기
    current_index = None
    for i, ch in enumerate(all_chapters):
        if ch.id == chapter.id:
            current_index = i
            break
    
    # 앞뒤 5개씩 슬라이싱 (현재 포함해서 총 11개)
    if current_index is not None:
        start_idx = max(0, current_index - 5)
        end_idx = min(len(all_chapters), current_index + 6)
        nearby_chapters = all_chapters[start_idx:end_idx]
    else:
        nearby_chapters = all_chapters[:11]
    
    # 연결된 BBS 게시글
    linked_posts = ChapterPost.objects.filter(
        chapter=chapter
    ).select_related('post', 'post__author', 'post__type')

    # 제목 밑에 본문 앞부분을 보여 준다 (나의 질의응답과 같은 방식)
    for lp in linked_posts:
        lp.preview = _qa_preview(lp.post.content)
        lp.body = _qa_body(lp.post.content)

    # 이 목차에 이어 붙인 주제별 문제집 (기출문제풀이)
    exam_sets = _chapter_exam_sets(chapter)

    return render(request, 'practice/chapter_detail.html', {
        'chapter': chapter,
        'content': content,
        'questions': questions,
        'siblings': nearby_chapters,  # 이름은 그대로 유지 (템플릿 호환)
        'linked_posts': linked_posts,
        'exam_sets': exam_sets,
    })



@login_required
def content_create(request):
    """학습 컨텐츠 작성"""
    books = Book.objects.all()
    
    # URL 파라미터로 chapter_id가 전달되면 해당 목차 선택
    chapter_id = request.GET.get('chapter_id')
    selected_chapter = None
    if chapter_id:
        selected_chapter = get_object_or_404(Chapter, id=chapter_id)
    
    if request.method == 'POST':
        chapter_id = request.POST.get('chapter_id')
        content_text = request.POST.get('content', '')
        
        if not chapter_id:
            messages.error(request, '목차를 선택해주세요.')
            return redirect('practice:content_create')
        
        chapter = get_object_or_404(Chapter, id=chapter_id)
        
        # 이미 컨텐츠가 있는지 확인
        if hasattr(chapter, 'content'):
            messages.error(request, '이 목차에는 이미 컨텐츠가 있습니다. 수정 기능을 사용하세요.')
            return redirect('practice:chapter_detail', chapter_id=chapter.id)
        
        ChapterContent.objects.create(
            chapter=chapter,
            content=content_text,
            author=request.user,
        )
        
        # Calculate nearby chapters for sidebar
        def natural_sort_key(ch):
            if not ch.code:
                return (999,)
            parts = ch.code.split('.')
            result = []
            for p in parts:
                try:
                    result.append(int(p))
                except ValueError:
                    result.append(999)
            return tuple(result)
        
        all_chapters = list(Chapter.objects.filter(book=chapter.book))
        all_chapters.sort(key=natural_sort_key)
        
        current_index = None
        for i, ch in enumerate(all_chapters):
            if ch.id == chapter.id:
                current_index = i
                break
        
        if current_index is not None:
            start_idx = max(0, current_index - 5)
            end_idx = min(len(all_chapters), current_index + 6)
            nearby_chapters = all_chapters[start_idx:end_idx]
        else:
            nearby_chapters = all_chapters[:11]
        
        # 연결된 BBS 게시글
        linked_posts = ChapterPost.objects.filter(
            chapter=chapter
        ).select_related('post', 'post__author', 'post__type')

        # 상세 화면과 같은 틀을 쓰므로 앞부분·본문도 함께 채운다.
        # 빠뜨리면 저장 직후 화면에서만 미리보기가 사라진다
        for lp in linked_posts:
            lp.preview = _qa_preview(lp.post.content)
            lp.body = _qa_body(lp.post.content)

        # 같은 페이지에서 저장 완료 표시 (다른 페이지로 메시지 전파 방지)
        return render(request, 'practice/chapter_detail.html', {
            'chapter': chapter,
            'content': chapter.content,
            'questions': PracticeQuestion.objects.filter(chapter=chapter).order_by('number'),
            'siblings': nearby_chapters,
            'saved': True,
            'linked_posts': linked_posts,
        })
    
    return render(request, 'practice/content_form.html', {
        'books': books,
        'selected_chapter': selected_chapter,
        'title': '컨텐츠 작성',
        'btn_text': '저장하기',
    })


@login_required
def content_update(request, content_id):
    """학습 컨텐츠 수정"""
    content = get_object_or_404(ChapterContent, id=content_id)
    
    # 작성자 또는 관리자만 수정 가능
    if content.author != request.user and not request.user.is_staff:
        messages.error(request, '수정 권한이 없습니다.')
        return redirect('practice:chapter_detail', chapter_id=content.chapter.id)
    
    if request.method == 'POST':
        content_text = request.POST.get('content', '')
        content.content = content_text
        content.save()
        
        # Calculate nearby chapters for sidebar
        def natural_sort_key(ch):
            if not ch.code:
                return (999,)
            parts = ch.code.split('.')
            result = []
            for p in parts:
                try:
                    result.append(int(p))
                except ValueError:
                    result.append(999)
            return tuple(result)
        
        all_chapters = list(Chapter.objects.filter(book=content.chapter.book))
        all_chapters.sort(key=natural_sort_key)
        
        current_index = None
        for i, ch in enumerate(all_chapters):
            if ch.id == content.chapter.id:
                current_index = i
                break
        
        if current_index is not None:
            start_idx = max(0, current_index - 5)
            end_idx = min(len(all_chapters), current_index + 6)
            nearby_chapters = all_chapters[start_idx:end_idx]
        else:
            nearby_chapters = all_chapters[:11]
        
        # 연결된 BBS 게시글
        linked_posts = ChapterPost.objects.filter(
            chapter=content.chapter
        ).select_related('post', 'post__author', 'post__type')

        # 상세 화면과 같은 틀을 쓰므로 앞부분·본문도 함께 채운다.
        # 빠뜨리면 수정 직후 화면에서만 미리보기가 사라진다
        for lp in linked_posts:
            lp.preview = _qa_preview(lp.post.content)
            lp.body = _qa_body(lp.post.content)

        # 같은 페이지에서 수정 완료 표시
        return render(request, 'practice/chapter_detail.html', {
            'chapter': content.chapter,
            'content': content,
            'questions': PracticeQuestion.objects.filter(chapter=content.chapter).order_by('number'),
            'siblings': nearby_chapters,
            'saved': True,
            'linked_posts': linked_posts,
        })
    
    return render(request, 'practice/content_form.html', {
        'content': content,
        'selected_chapter': content.chapter,
        'title': '컨텐츠 수정',
        'btn_text': '수정하기',
        'is_edit': True,
    })


@login_required
@staff_member_required
def api_search_bbs_posts(request, chapter_id):
    """BBS 게시글 검색/목록 API"""
    chapter = get_object_or_404(Chapter, id=chapter_id)
    query = request.GET.get('q', '').strip()
    offset = int(request.GET.get('offset', 0))
    limit = 10

    from bbs.models import Post

    linked_ids = set(
        ChapterPost.objects.filter(chapter=chapter).values_list('post_id', flat=True)
    )

    # 한 글은 목차 하나에만 붙인다. 다른 목차에 이미 붙어 있으면
    # 어느 목차인지 함께 알려 주어 연결을 막는다
    elsewhere = {}
    for cp in (ChapterPost.objects
               .exclude(chapter=chapter)
               .select_related('chapter')):
        if cp.post_id in elsewhere:
            continue
        ch = cp.chapter
        elsewhere[cp.post_id] = ('%s %s' % (ch.code or '', ch.title or '')).strip()

    if query:
        posts_qs = Post.objects.filter(
            Q(title__icontains=query) | Q(content__icontains=query)
        )
    else:
        posts_qs = Post.objects.all()

    posts = posts_qs.select_related(
        'author', 'type'
    ).order_by('-created_at')[offset:offset + limit]

    has_more = posts_qs.count() > offset + limit

    posts_data = []
    for post in posts:
        posts_data.append({
            'id': post.pk,
            'title': post.title,
            'type': post.type.name if post.type else None,
            'author': post.author.first_name or post.author.username,
            'created_at': post.created_at.strftime('%Y-%m-%d'),
            'hits': post.hits,
            'already_linked': post.pk in linked_ids,
            'linked_elsewhere': elsewhere.get(post.pk),
        })

    return JsonResponse({'posts': posts_data, 'has_more': has_more})


@login_required
@staff_member_required
def api_link_post(request, chapter_id):
    """게시글을 목차에 연결"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    chapter = get_object_or_404(Chapter, id=chapter_id)
    post_id = request.POST.get('post_id')

    if not post_id:
        return JsonResponse({'success': False, 'error': '게시글 ID가 없습니다.'})

    from bbs.models import Post
    post = get_object_or_404(Post, id=post_id)

    if ChapterPost.objects.filter(chapter=chapter, post=post).exists():
        return JsonResponse({'success': False, 'error': '이미 연결된 게시글입니다.'})

    other = (ChapterPost.objects
             .exclude(chapter=chapter)
             .filter(post=post)
             .select_related('chapter')
             .first())
    if other:
        ch = other.chapter
        return JsonResponse({
            'success': False,
            'error': '이미 다른 목차에 연결된 글입니다 (%s %s).'
                     % (ch.code or '', ch.title or ''),
        })

    ChapterPost.objects.create(
        chapter=chapter,
        post=post,
        linked_by=request.user,
    )

    return JsonResponse({
        'success': True,
        'link': {
            'post_id': post.pk,
            'title': post.title,
            'type': post.type.name if post.type else None,
            'author': post.author.first_name or post.author.username,
            'created_at': post.created_at.strftime('%Y-%m-%d'),
            'hits': post.hits,
            # 목록에 바로 붙일 수 있게 앞부분과 본문도 함께 보낸다
            'preview': str(_qa_preview(post.content)),
            'content': _qa_body(post.content),
        }
    })


@login_required
@staff_member_required
def api_unlink_post(request, chapter_id):
    """게시글 연결 해제"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    chapter = get_object_or_404(Chapter, id=chapter_id)
    post_id = request.POST.get('post_id')

    if not post_id:
        return JsonResponse({'success': False, 'error': '게시글 ID가 없습니다.'})

    deleted, _ = ChapterPost.objects.filter(
        chapter=chapter, post_id=post_id
    ).delete()

    if deleted == 0:
        return JsonResponse({'success': False, 'error': '연결 정보를 찾을 수 없습니다.'})

    return JsonResponse({'success': True, 'post_id': int(post_id)})

